import os
import io
import jwt
import datetime
import pandas as pd
from flask import jsonify, request
from supabase import create_client
from openpyxl import load_workbook

SUPABASE_STAFF_URL = os.environ.get("SUPABASE_STAFF_URL")
SUPABASE_STAFF_KEY = os.environ.get("SUPABASE_STAFF_KEY")
JWT_SECRET = os.environ.get("JWT_SECRET")

supabase_staff = create_client(SUPABASE_STAFF_URL, SUPABASE_STAFF_KEY)

MASTER_PATH = os.path.join(os.path.dirname(__file__), "スタッフマスター.xlsx")

# 勤務日数→%達成・%維持テーブル（スプレッドシートより）
RATE_TABLE = {
    5: {23: (None, None), 22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20), 16: (2.60, 2.20),
        15: (2.60, 2.20), 14: (2.60, 2.20), 13: (2.75, 2.30), 12: (2.75, 2.30),
        11: (2.75, 2.30), 10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
    4: {23: (None, None), 22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20), 16: (2.60, 2.20),
        15: (2.60, 2.20), 14: (2.60, 2.20), 13: (2.75, 2.30), 12: (2.75, 2.30),
        11: (2.75, 2.30), 10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
    3: {23: (None, None), 22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20), 16: (2.60, 2.20),
        15: (2.60, 2.20), 14: (2.60, 2.20), 13: (2.75, 2.30), 12: (2.75, 2.30),
        11: (2.75, 2.30), 10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
}

def load_staff_master():
    wb = load_workbook(MASTER_PATH, data_only=True)
    ws1 = wb["スタッフマスター"]
    ws2 = wb["時給マスター"]

    master = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        staff_id, name, site, rank = row[0], row[1], row[2], row[3]
        if staff_id and rank:
            master[str(staff_id).strip()] = {
                "name": name, "site": site, "rank": rank,
                "hourly_wage": 0, "mgmt_fee": 0, "work_pattern": 5
            }

    for row in ws2.iter_rows(min_row=2, values_only=True):
        staff_id, name, wage, note = row[0], row[1], row[2], row[3]
        if staff_id and str(staff_id).strip() in master:
            sid = str(staff_id).strip()
            master[sid]["hourly_wage"] = int(wage) if wage else 0
            master[sid]["mgmt_fee"] = 3030 if note and "管理料" in str(note) else 0

    return master

def register_staff_routes(app):

    @app.route("/health_staff")
    def health_staff():
        return jsonify({"status": "ok", "service": "staff-dashboard"})

    @app.route("/staff/summary")
    def staff_summary():
        try:
            month = request.args.get("month")
            if not month:
                return jsonify({"error": "monthパラメータが必要です"}), 400

            target_month = month + "-01"

            # スタッフマスター読み込み
            master = load_staff_master()

            # アポイントメント取得
            apo_res = supabase_staff.table("appointments")\
                .select("*")\
                .eq("target_month", target_month)\
                .execute()
            apo_rows = apo_res.data

            # 勤怠取得
            att_res = supabase_staff.table("attendance")\
                .select("*")\
                .eq("target_month", target_month)\
                .execute()
            att_rows = att_res.data

            # スタッフ別集計
            results = {}
            for sid, info in master.items():
                results[sid] = {
                    "staff_id": sid,
                    "name": info["name"],
                    "site": info["site"],
                    "rank": info["rank"],
                    "apo_amount": 0,
                    "cxl_amount": 0,
                    "fb_amount": 0,
                    "sales": 0,
                    "work_days": 0,
                    "target_achieve": 0,
                    "target_maintain": 0,
                    "achieve_rate": 0,
                }

            # アポ・CXL・FB集計
            for row in apo_rows:
                sid = row["staff_id"]
                if sid not in results:
                    continue
                cancel = row.get("cancel_date") or ""
                # 考慮キャンセル除外
                if "考慮" in str(cancel):
                    results[sid]["cxl_amount"] += row.get("amount", 0)
                    continue
                if cancel and cancel not in ["None", ""]:
                    results[sid]["cxl_amount"] += row.get("amount", 0)
                else:
                    results[sid]["apo_amount"] += row.get("amount", 0)
                results[sid]["fb_amount"] += row.get("fb_amount", 0)

            # 勤怠集計
            for row in att_rows:
                sid = row["staff_id"]
                if sid not in results:
                    continue
                if (row.get("work_hours") or 0) > 0:
                    results[sid]["work_days"] += 1

            # 売上・目標・達成率計算
            for sid, r in results.items():
                info = master[sid]
                r["sales"] = r["apo_amount"] - r["cxl_amount"] + r["fb_amount"]

                wage = info["hourly_wage"]
                mgmt = info["mgmt_fee"]
                pattern = info["work_pattern"]
                days = r["work_days"]

                rate_row = RATE_TABLE.get(pattern, {}).get(days)
                if rate_row and rate_row[0]:
                    base = wage * 8 + 1000 + mgmt
                    r["target_achieve"] = int(base * days * rate_row[0])
                    r["target_maintain"] = int(base * days * rate_row[1])
                    if r["target_achieve"] > 0:
                        r["achieve_rate"] = round(r["sales"] / r["target_achieve"] * 100, 1)

            return jsonify({"status": "ok", "data": list(results.values())})

        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/appointments_json", methods=["POST"])
    def upload_appointments_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["acquired_date"][:7] + "-01" if r.get("acquired_date") else None
            supabase_staff.table("appointments").upsert(records, on_conflict="appointment_id").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/productivity_json", methods=["POST"])
    def upload_productivity_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["call_date"][:7] + "-01" if r.get("call_date") else None
            supabase_staff.table("productivity").upsert(records, on_conflict="staff_id,call_date").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/attendance_json", methods=["POST"])
    def upload_attendance_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["work_date"][:7] + "-01" if r.get("work_date") else None
            supabase_staff.table("attendance").upsert(records, on_conflict="staff_id,work_date").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
